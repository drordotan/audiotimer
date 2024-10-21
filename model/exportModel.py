
import openpyxl
from openpyxl.styles import numbers

class ExportModel:

    # Define the headers based on the dictionary keys
    HEADERS = ["subject", "trail_num", "n1", "n2", "result",
                "start", "end", "1 start", "1 end", "2 start", "2 end",
                "final start", "final end"]

    def __init__(self) -> None:
        pass


    def save_to_xlsx(self, dict_list, filename):
        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers to the first row of the sheet
        for col_num, header in enumerate(self.HEADERS, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write the data from the list of dictionaries
        for row_num, item in enumerate(dict_list, 2):
            for col_num, header in enumerate(self.HEADERS, 1):
                cell_value = item.get(header, "")
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                
                # Apply text format to cells with time-like values
                if isinstance(cell_value, str) and ":" in cell_value:
                    cell.number_format = numbers.FORMAT_TEXT

        # Save the workbook to the specified filename
        workbook.save(filename)

